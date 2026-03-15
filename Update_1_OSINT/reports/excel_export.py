"""
excel_export.py
---------------
Generates Excel reports for OSINT / SOC dashboard
Pylance-clean, merged-cell safe, production ready
"""

import os
import sqlite3
from datetime import datetime

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from data.db_config import DB_PATH


# ================= PATHS =================
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
REPORT_DIR = os.path.join(BASE_DIR, "reports", "generated")
os.makedirs(REPORT_DIR, exist_ok=True)


# ================= MAIN FUNCTION =================
def generate_excel_report() -> str:
    """
    Export analysis logs into Excel
    Returns generated file path
    """

    # ---------- DATABASE ----------
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()

    cur.execute("""
        SELECT
            target,
            threat_level,
            threat_type,
            risk_score,
            vt_score,
            abuse_score,
            open_ports,
            created_at
        FROM analysis_logs
        ORDER BY created_at DESC
    """)

    rows = cur.fetchall()
    conn.close()

    # ---------- WORKBOOK ----------
    wb = Workbook()
    ws = wb.active
    assert isinstance(ws, Worksheet)

    ws.title = "OSINT Threat Report"

    # ---------- HEADERS ----------
    headers = [
        "IP / Domain",
        "Threat Level",
        "Threat Type",
        "Risk Score",
        "VirusTotal Score",
        "AbuseIPDB Score",
        "Open Ports",
        "Scan Time (UTC)"
    ]

    ws.append(headers)

    # ---------- HEADER STYLE ----------
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for col_index in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_index)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # ---------- DATA ----------
    for row_idx, row in enumerate(rows, start=2):
        ws.append(row)
        threat_level = str(row[1]).lower()
        
        # Color coding rows based on threat level
        if "high" in threat_level or "critical" in threat_level:
            for col_idx in range(1, len(row) + 1):
                ws.cell(row=row_idx, column=col_idx).font = Font(color="EF4444", bold=True)
        elif "medium" in threat_level:
            for col_idx in range(1, len(row) + 1):
                ws.cell(row=row_idx, column=col_idx).font = Font(color="F59E0B")

    # ---------- AUTO COLUMN WIDTH (MERGED-CELL SAFE) ----------
    for col_index in range(1, ws.max_column + 1):
        max_length = 0

        for row_index in range(1, ws.max_row + 1):
            value = ws.cell(row=row_index, column=col_index).value
            if value is not None:
                max_length = max(max_length, len(str(value)))

        column_letter = get_column_letter(col_index)
        ws.column_dimensions[column_letter].width = max_length + 3

    # ---------- SAVE ----------
    filename = f"osint_report_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    file_path = os.path.join(REPORT_DIR, filename)

    wb.save(file_path)

    return file_path
